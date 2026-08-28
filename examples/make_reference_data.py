"""
make_reference_data.py — generate examples/reference_mesoporous.xlsx
===================================================================

Synthetic, committable reference dataset for BET_analyser: a Type IV(a)
mesoporous isotherm (BET multilayer + capillary condensation + hysteresis
loop) with a well-defined BJH mesopore peak, written in the exact 4-sheet
layout that ``bet_analysis.read_bet_xls`` expects.

Single source of truth (constants at the top of this file):

    C          = 120.0                 BET C constant
    n_m                              monolayer capacity — NOT chosen freely;
                                     derived by fixed-point iteration so that
                                     S_BET = n_m × 4.353 equals S_BJH (the
                                     reference material is purely mesoporous,
                                     so the two areas must agree).
    S_BET_true = n_m × 4.353          m² g⁻¹  (reported after convergence)

The low-pressure branch is the finite-layer (BDDT) BET equation — reused from
``tests/synthetic_isotherms.bet`` — with the multilayer carried on *top* of a
capillary-condensation sigmoid, i.e. the single continuous form

    Va(x) = bet(x, n_m, C, n=N_LAYERS) + step(x, X_STEP, STEP_W, STEP_H)

This has no step discontinuity and keeps the multilayer growing (rather than
freezing it at the condensation onset). ``bet`` is called with a finite number
of layers so the isotherm saturates to a genuine Type IV plateau; with the
default ``n=None`` the infinite-layer BET diverges as p/p0 → 1, which would
break both the plateau and the fixed-point iteration below. ``N_LAYERS = 7``
keeps the BET window (0.05–0.35) within ~0.3 % of the infinite-layer form, so
the fitted BET area still equals ``n_m × 4.353``.

Every sheet derives from ``n_m``, ``C`` and the generated ads/des arrays —
nothing is hardcoded to make the answer "come out right":

  * AdsDes cols 5,6           : the generated p/p0 and Va.
  * BET col 2                 : y = 1/[Va(1/p/p0 − 1)] from the SAME Va values
                                written into AdsDes.
  * Summary "Vm" / "as,BET"   : n_m / n_m × 4.353.
  * Summary "C"               : the C used to generate the curve.
  * Summary "Total pore volume(p/p0=0.990)" : Gurvich rule, from the Va nearest
                                p/p0 = 0.990, using ``bet_analysis.N2_STP_TO_LIQUID``.
  * Summary "rp,peak(Area)"   : radius at the maximum of the BJH dVp/drp col.
  * Summary "ap" / "Vp"       : BJH cumulative surface area / pore volume,
                                running integrals of the dVp/drp column.
  * Summary "Average pore diameter" : 4 × Vp / S_BJH (cylindrical model).

Instrument metadata rows that cannot be derived (sample/operator/date/
instrument) are filled with the placeholder ``SYNTHETIC-REFERENCE``; the sample
name is ``SYNTHETIC-mesoporous-reference`` so the file cannot be mistaken for
measured data.

Deterministic: running it again reproduces the file identically (no RNG).
No new runtime dependencies (openpyxl is already in requirements.txt).
"""
import io
import os
import re
import sys
import zipfile
from datetime import datetime

import numpy as np
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
sys.path.insert(0, REPO)
sys.path.insert(0, os.path.join(REPO, "tests"))

# Reuse the project's own constants and physics generators (no second BET
# implementation, no duplicated physical constants).
from bet_analysis import N2_BET_FACTOR, N2_STP_TO_LIQUID  # noqa: E402
from synthetic_isotherms import bet, step, desorption_closed  # noqa: E402

_trapezoid = getattr(np, "trapezoid", None) or np.trapz

# ══════════════════════════════════════════════════════════════
# Single source of truth
# ══════════════════════════════════════════════════════════════
C        = 120.0                 # BET C constant
N_M_INIT = 90.0                  # initial guess for n_m (fixed-point iteration)
N_LAYERS = 7                     # finite-layer BET: saturating multilayer, and
                                 # keeps the BET window within ~0.3 % of the
                                 # infinite-layer form (see module docstring).

SAMPLE      = "SYNTHETIC-mesoporous-reference"
PLACEHOLDER = "SYNTHETIC-REFERENCE"

# Isotherm shape: finite-layer BET multilayer + capillary-condensation sigmoid
# (single continuous form, no freeze/discontinuity).
X_STEP = 0.60
STEP_W = 0.08
STEP_H = 200.0                   # cm³(STP) g⁻¹ added by capillary condensation

# Desorption branch (H1 loop, closed at both ends): a volume-domain bump
# vd = v_ads + AMPLITUDE·sin(pi·u) that vanishes at CLOSE_AT and p/p0 = 1
# (desorption_closed). AMPLITUDE = 25 keeps the closure gap < 0.02, the branch
# monotonic, slope_ratio_max <= 1.10, and a visible H1 loop (area_norm ~ 0.024).
AMPLITUDE = 25.0
CLOSE_AT  = 0.45

# BJH mesopore peak (radius, nm) → 12 nm diameter peak.
R_PEAK  = 6.0
R_SIGMA = 0.35

BET_CUTOFF = 0.50                # BET sheet includes adsorption points below this p/p0

OUT = os.path.join(HERE, "reference_mesoporous.xlsx")


def _grid():
    """p/p0 grid: dense in the BET region, 0.05–0.35, and up to 0.995."""
    lo = np.logspace(np.log10(5e-4), np.log10(0.05), 16, endpoint=False)
    mid = np.linspace(0.05, 0.35, 14, endpoint=False)
    hi = np.linspace(0.35, 0.995, 30, endpoint=True)
    return np.unique(np.concatenate([lo, mid, hi]))


def make_isotherm(n_m):
    x = _grid()
    va = bet(x, n_m, C, n=N_LAYERS) + step(x, X_STEP, STEP_W, STEP_H)
    ads = np.column_stack([x, va])
    des = desorption_closed(x, va, AMPLITUDE, CLOSE_AT)
    return x, ads, des


def make_bjh(vp_total):
    """BJH pore-size distribution whose total volume equals the Gurvich Vp.

    Returns (rp, dV, cum_vp, cum_sap, rp_peak, s_bjh, vp_bjh). ``dV`` is the
    differential dVp/drp (per radius), and the cumulative columns are running
    trapezoidal integrals of ``dV`` over the same radius grid (cylindrical
    pore geometry: dS = 2 dV / r, with 1 cm³→1e-6 m³ and 1 nm→1e-9 m).
    """
    rp = np.logspace(np.log10(1.0), np.log10(25.0), 60)
    dV = np.exp(-0.5 * ((np.log(rp) - np.log(R_PEAK)) / R_SIGMA) ** 2)
    dV /= _trapezoid(dV, rp)
    dV *= vp_total

    drp = np.diff(rp)
    dV_seg = 0.5 * (dV[:-1] + dV[1:]) * drp
    cum_vp = np.concatenate([[0.0], np.cumsum(dV_seg)])

    rp_mid = 0.5 * (rp[:-1] + rp[1:])
    dS_seg = 2.0e3 * dV_seg / rp_mid
    cum_sap = np.concatenate([[0.0], np.cumsum(dS_seg)])

    rp_peak = rp[int(np.argmax(dV))]
    s_bjh = float(cum_sap[-1])
    vp_bjh = float(cum_vp[-1])
    return rp, dV, cum_vp, cum_sap, rp_peak, s_bjh, vp_bjh


def write_workbook(ads, des, bet_pts, bjh_cols, summary):
    rp, dV, cum_vp, cum_sap = bjh_cols

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("AdsDes")
    ws.append(["ADS", None, None, None, None, "p/p0", "Va (cm3/g STP)"])
    for p, v in ads:
        ws.append([None, None, None, None, None, p, v])
    ws.append(["DES", None, None, None, None, "p/p0", "Va (cm3/g STP)"])
    for p, v in des[::-1]:  # desorption written high → low pressure
        ws.append([None, None, None, None, None, p, v])

    ws = wb.create_sheet("BET")
    ws.append(["No", "p/p0", "y", "idx"])
    for i, (p, y) in enumerate(bet_pts):
        ws.append([None, p, y, i])
    ws.append(["Starting point", "START", None, summary["start_pt"]])
    ws.append(["End point", "END", None, summary["end_pt"]])

    ws = wb.create_sheet("BJH")
    ws.append(["No", None, "rp/nm", "dVp/drp", "cum Vp", "cum Sap"])
    for row in zip(rp, dV, cum_vp, cum_sap):
        ws.append([None, None, *row])

    ws = wb.create_sheet("Summary")
    ws.append(["Sample", None, None, SAMPLE])
    ws.append(["Instrument", None, None, PLACEHOLDER])
    ws.append(["Operator", None, None, PLACEHOLDER])
    ws.append(["Date", None, None, PLACEHOLDER])
    ws.append(["Vm", None, None, summary["Vm"]])
    ws.append(["as,BET", None, None, summary["S_BET"]])
    ws.append(["C", None, None, summary["C"]])
    ws.append(["Total pore volume(p/p0=0.990)", None, None, summary["Vp_total"]])
    ws.append(["Average pore diameter", None, None, summary["dp_avg"]])
    ws.append(["rp,peak(Area)", None, None, summary["rp_peak_BJH"]])
    ws.append(["ap", None, None, summary["S_BJH"]])
    ws.append(["Vp", None, None, summary["Vp_BJH"]])

    _save_deterministic(wb, OUT)


_FIXED_TS = (2026, 1, 1, 0, 0, 0)  # fixed zip-entry + core-property timestamp
_FIXED_DT = "2026-01-01T00:00:00Z"
_FIX_MODIFIED_RE = re.compile(
    r"(<dcterms:modified[^>]*>)([^<]*)(</dcterms:modified>)"
)


def _save_deterministic(wb, path):
    """Save the workbook so that re-running the script reproduces it exactly.

    openpyxl stamps ``docProps/core.xml`` (created/modified) and every zip
    entry with the current time, which makes two runs differ byte-for-byte.
    Pin both to a fixed value so the generated file is reproducible.
    """
    wb.properties.creator = "BET_analyser"
    wb.properties.created = datetime(*_FIXED_TS)
    wb.properties.modified = datetime(*_FIXED_TS)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with zipfile.ZipFile(buf, "r") as zin, \
            zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "docProps/core.xml":
                # openpyxl forces dcterms:modified to "now" inside save(); pin it.
                data = _FIX_MODIFIED_RE.sub(
                    lambda m: m.group(1) + _FIXED_DT + m.group(3),
                    data.decode("utf-8")).encode("utf-8")
            zi = zipfile.ZipInfo(item.filename, date_time=_FIXED_TS)
            zi.compress_type = zipfile.ZIP_DEFLATED
            zi.external_attr = item.external_attr
            zi.create_system = item.create_system
            zout.writestr(zi, data)


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

    # ── B2: fixed-point iteration ──────────────────────────────
    # n_m is not chosen freely: iterate until S_BET (= n_m × 4.353) agrees
    # with S_BJH (the surface area implied by the BJH pore geometry, itself
    # normalised to the isotherm's Gurvich pore volume). This closes the loop
    # that previously left the two areas 2.3× apart.
    n_m = N_M_INIT
    for it in range(20):
        x, ads, des = make_isotherm(n_m)

        i099 = int(np.argmin(np.abs(ads[:, 0] - 0.990)))
        vp_total = ads[i099, 1] * N2_STP_TO_LIQUID

        rp, dV, cum_vp, cum_sap, rp_peak, s_bjh, vp_bjh = make_bjh(vp_total)

        s_bet = n_m * N2_BET_FACTOR
        ratio = s_bet / s_bjh
        if abs(ratio - 1.0) < 0.02:
            break
        n_m = s_bjh / N2_BET_FACTOR
    else:
        raise RuntimeError("B2 fixed-point iteration did not converge in 20 "
                           "iterations")

    # ── B1: smoothness check (no step discontinuity) ───────────
    diff_va = np.abs(np.diff(ads[:, 1]))
    step_profile = step(x, X_STEP, STEP_W, STEP_H)
    cond = (step_profile[1:] > 0.05 * STEP_H) & (step_profile[1:] < 0.95 * STEP_H)
    max_diff = float(diff_va.max())
    med_cond = float(np.median(diff_va[cond]))
    assert max_diff <= 3.0 * med_cond, "isotherm has a step discontinuity"

    # BET plot points: every adsorption point below BET_CUTOFF, with
    # y = 1/[Va(1/x - 1)] computed from the SAME Va written into AdsDes.
    mask = ads[:, 0] < BET_CUTOFF
    x_bet = ads[mask, 0]
    v_bet = ads[mask, 1]
    y_bet = 1.0 / (v_bet * (1.0 / x_bet - 1.0))
    bet_pts = np.column_stack([x_bet, y_bet])

    in_range = (x_bet >= 0.05) & (x_bet <= 0.35)
    assert in_range.sum() >= 5, "ISO 9277 requires >= 5 points in 0.05–0.35"
    start_pt = int(np.where(x_bet >= 0.05)[0][0])
    end_pt = int(np.where(x_bet <= 0.35)[0][-1])

    dp_avg = 4.0e3 * vp_bjh / s_bjh  # nm, cylindrical 4 V / S

    summary = {
        "Vm": n_m,
        "S_BET": s_bet,
        "C": C,
        "Vp_total": vp_total,
        "dp_avg": dp_avg,
        "rp_peak_BJH": rp_peak,
        "S_BJH": s_bjh,
        "Vp_BJH": vp_bjh,
        "start_pt": start_pt,
        "end_pt": end_pt,
    }

    write_workbook(ads, des, bet_pts, (rp, dV, cum_vp, cum_sap), summary)

    # ── B1/B2 acceptance + round-trip report ───────────────────
    print("=== B1 (smoothness) ===")
    print(f"max|diff(va)|                    = {max_diff:.4f}")
    print(f"median|diff| over condensation   = {med_cond:.4f}")
    print(f"3 x median                      = {3.0 * med_cond:.4f}")
    print(f"pass (max <= 3x median)         = {max_diff <= 3.0 * med_cond}")

    print("=== B2 (self-consistency) ===")
    print(f"iterations                      = {it + 1}")
    print(f"n_m (final)                     = {n_m:.6f} cm³(STP) g⁻¹")
    print(f"C                               = {C}")
    print(f"S_BET_true = n_m × 4.353        = {s_bet:.4f} m²/g")
    print(f"S_BJH                           = {s_bjh:.4f} m²/g")
    print(f"S_BET / S_BJH                   = {ratio:.4f}")
    print(f"pass (|ratio - 1| <= 0.15)      = {abs(ratio - 1.0) <= 0.15}")

    # Round-trip check: read the file back with the project's own reader and
    # confirm every value survived the pandas/openpyxl positional read.
    from bet_analysis import read_bet_xls, verify_bet

    data = read_bet_xls(OUT)
    print("=== round-trip (read_bet_xls) ===")
    print("summary:", data["summary"])
    for k in ("ads", "des", "bet_pts", "bjh"):
        print(f"{k}: shape = {tuple(data[k].shape)}")

    res = verify_bet(data["bet_pts"], data["summary"])
    print("=== verify_bet ===")
    print(f"S_BET_calc   = {res['S_BET_calc']:.4f} m²/g  (summary S_BET = {s_bet:.4f})")
    print(f"Vm (calc)    = {res['Vm']:.4f}  (n_m = {n_m:.4f})")
    print(f"C  (calc)    = {res['C']:.2f}  (C   = {C})")
    print(f"R²           = {res['R2']:.6f}")
    print(f"BET window   = points {summary['start_pt']}–{summary['end_pt']} "
          f"({end_pt - start_pt + 1} points)")
    print(f"Wrote: {OUT}")


if __name__ == "__main__":
    main()
